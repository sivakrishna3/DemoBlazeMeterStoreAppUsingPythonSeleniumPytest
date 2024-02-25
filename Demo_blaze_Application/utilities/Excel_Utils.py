import openpyxl
import openpyxl.reader.excel


def get_row_count(sheet_name):
    Excel_file_path = ".\\Test_Data\\Test_Data_for_DDT.xlsx"
    try:
        workbook = openpyxl.load_workbook(Excel_file_path)
        Sheet = workbook[sheet_name]
        return Sheet.max_row
    except FileNotFoundError:
        print(f"Error: File not found - {Excel_file_path}")
    except KeyError:
        print(f"Error: Sheet not found - {sheet_name}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[sheet_name]
    return Sheet.max_column


def read_data():
    global sheet
    final_list = []
    Excel_file_path = ".\\Test_Data\\Test_Data_for_DDT.xlsx"
    try:
        workbook = openpyxl.load_workbook(Excel_file_path)
        sheet = workbook["Login_Data"]
        total_rows = sheet.max_row
        total_columns = sheet.max_column
        for row in range(2, total_rows+1):
            row_list = []
            for col in range(2, total_columns+1):
                row_list.append(sheet.cell(row=row, column=col).value)
            final_list.append(row_list)
        return final_list
    except FileNotFoundError:
        print(f"Error: File not found - {Excel_file_path}")
    except KeyError:
        print(f"Error: Sheet not found - {sheet}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def write_data(file, sheet_name, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[sheet_name]
    Sheet.cell(row=row_num, column=column_num).value = data
    workbook.save(file)
